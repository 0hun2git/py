from openpyxl import Workbook
from openpyxl import load_workbook
from random import *

#wb = Workbook() # wb 생성
wb = load_workbook("rpa.xlsx")
ws = wb.active # wb에 활성화 된 ws 선택

# 값 읽기
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

# Sheet
ws.title = "0hun1" # ws 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # ws 색상변경
ws2 = wb.create_sheet("0hun2",1) # 두 번째 위치에 ws 생성 
ws1 = wb["0hun1"] # 0hun1 시트 선택
ws1["A1"] = "Test1" # 0hun1 A1셀에 Test 입력
ws1["A2"] = "Test2"
ws1["A3"] = "Test3"
ws1["B1"] = "Test4"
ws1["B2"] = "Test5"
ws1["B3"] = "Test6"
print(wb.sheetnames)
ws3 = wb.copy_worksheet(ws1) #
ws3.title = "Copied 0hun1"

# cell(row, column)
print(ws1["A1"])
print(ws1["A1"].value)
print(ws1.cell(row=1, column=1).value)
c = ws1.cell(row=1, column=3, value=10)
print(c.value)
print(ws1["A10"].value)

# 값 쓰기
index=1
for x in range(1, 11):
    for y in range(1, 11):
        #ws.cell(row=x, column=y, value=randint(0,100))
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("rpa.xlsx")
wb.close()